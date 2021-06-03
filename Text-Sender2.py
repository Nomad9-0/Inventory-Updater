
import openpyxl
from openpyxl import Workbook
import os
import twilio
from twilio.rest import Client
from datetime import date
from twilio.base.exceptions import TwilioRestException
import logging

excel_file = openpyxl.load_workbook('D:\\javie\\Desktop\\Python\\TextSender\\Final\\OOSL2.xlsx')
custy_list = excel_file['Main']
   
account_sid = os.environ['TWILIO_ACCOUNT_SID']
auth_token = os.environ['TWILIO_AUTH_TOKEN']
client = Client(account_sid, auth_token)

def write_data():
    user_choice = 'Y'
    while user_choice != 'N':
        custy_name = input('Enter the customer\'s name:\n')
        custy_phone = int(input('Enter the customer\'s 10-digit phone number (IMPORTANT: only numbers, no spaces):\n'))
        product_name = input('Enter the product they are looking for:\n')
        date_req = date.today().strftime("%m/%d/%Y")
        new_row = [custy_name, custy_phone, product_name, date_req, "No", "-", "No", "-"]
        custy_list.append(new_row)
        user_choice= input('Do you wish to input more info? Y/N\n').upper()
    
    excel_file.save('D:\\javie\\Desktop\\Python\\TextSender\\Final\\OOSL2.xlsx')   
    print('Done writing to Out of Stock List!')
    
def update_data():
    custy_name = input('For which customer do you want to update data for?\n')
    for col in custy_list.iter_cols(min_row=2, max_col=1,max_row=custy_list.max_row):
        for cell in col:
            if cell.value == custy_name:
                print('Match found')
                exit()
            else:
                print('No match')
    
def read_data():   
    i = 2
    def send_text():
        name = custy_name
        product = product_req
        msid = 'MG8863fcb642269f21de0daeac40806feb'
        msg = '%s, we have restocked on an item you were looking for. Come to The Roots and check it out! https://linktr.ee/therootsdispensary' % (str(name)) #creates message to send
        phoneNum = '+1' + str(custy_phone) #sets phone number to send text tomessage = client.messages.create(body= msg, messaging_service_sid = msid,to= phoneNum					 )
        print('Message sent successfully. \n')
        custy_list.cell(row=i, column=7).value = 'Yes' #updates Msg cell to reflect that text has been sent, preventing duplicates
        whats_Today = date.today().strftime("%m/%d/%Y")
        custy_list.cell(row=i, column=8).value = whats_Today
                            
	
    while i <= custy_list.max_row:
        custy_name = str(custy_list.cell(row = i, column = 1).value)
        custy_phone = str(custy_list.cell(row = i, column = 2).value)
        product_req = str(custy_list.cell(row = i, column = 3).value)
        date_req = str(custy_list.cell(row = i, column = 4).value)
        restocked = str(custy_list.cell(row = i, column = 5).value)
        date_restocked = str(custy_list.cell(row = i, column = 6).value)
        has_texted = str(custy_list.cell(row = i, column = 7).value)
        text_date = str(custy_list.cell(row = i, column = 8).value)
        
        print('%s called for %s on %s.' % (custy_name, product_req, date_req))
        if restocked == 'Yes': #checks if item is restocked
            print('We restocked %s on %s.' % (product_req, date_restocked))
            if has_texted == 'Yes': #checks if text message has been sent
                print('On %s, a text message was sent to %s at phone number %s. \n' 
                % (text_date, custy_name, custy_phone)) 
            elif has_texted == 'No':
                print('A text message has not yet been sent. \n')
                while True:
                    choice = input('Would you like to send a message? Y/N: \n').upper()
                    try:
                        if choice == 'Y':
                            send_text()
                            break
                        elif choice == 'N':
                            print('No text will be sent.')
                            break
                            
                        else:
                            print('Incorrect Input. Try again.')
                            continue
                    except TwilioRestException as e:
                        print('Error sending text.')
                        print(e)
                else:
                    print('Input not recognized. Please try again.\n')
        elif restocked == 'No':
            print('We have not restocked %s yet. Please check back\n' % (product_req))
        i += 1
        excel_file.save('D:\\javie\\Desktop\\Python\\TextSender\\Final\\OOSL2.xlsx')
        
    print('Done parsing through Out of Stock List!')

def main():
	while True:
		funct = input('Do you want to read, write, or update the list?\n').upper()
		if funct == 'READ':
			read_data()
			break
		elif funct == 'WRITE':
			write_data()
			break
		elif funct == 'UPDATE':
			update_data()
			break
		else:
			print('Incorrect input; try again')
			continue
	
main()	

